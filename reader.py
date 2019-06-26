#encoding=latin1

# POC tool to read Excel using Python
# Data will be used to create subtasks / add attachments to Jira main issues
# Created either via this tool or Excel import plugin
# 
#
#
# Author mika.nokka1@gmail.com 
#
#
#from __future__ import unicode_literals

import openpyxl 
import sys, logging
import argparse
#import re
from collections import defaultdict
from author import Authenticate  # no need to use as external command
from author import DoJIRAStuff
from CreateIssue import CreateIssue 
import glob
import json # for json dumo
from sqlalchemy.sql.expression import false
import re
import time
 
__version__ = "0.1.1396"


logging.basicConfig(level=logging.INFO) # IF calling from Groovy, this must be set logging level DEBUG in Groovy side order these to be written out



def main(argv):
    
    JIRASERVICE=""
    JIRAPROJECT=""
    PSWD=''
    USER=''
  
    logging.debug ("--Python starting Excel reading --") 

 
    parser = argparse.ArgumentParser(usage="""
    {1}    Version:{0}     -  mika.nokka1@gmail.com for Ambientia
    
    USAGE:
    -filepath  | -p <Path to Excel file directory>
    -filename   | -n <Excel filename>

    """.format(__version__,sys.argv[0]))

    parser.add_argument('-f','--filepath', help='<Path to Excel file directory>')
    parser.add_argument('-n','--filename', help='<Main tasks Excel filename>')
    parser.add_argument('-m','--subfilename', help='<Subtasks Excel filename>')
    parser.add_argument('-v','--version', help='<Version>', action='store_true')
    
    parser.add_argument('-w','--password', help='<JIRA password>')
    parser.add_argument('-u','--user', help='<JIRA user>')
    parser.add_argument('-s','--service', help='<JIRA service>')
    parser.add_argument('-p','--project', help='<JIRA project>')
   
    parser.add_argument('-a','--attachemnts', help='<Attachment directory>')
        
    args = parser.parse_args()
    
    if args.version:
        print 'Tool version: %s'  % __version__
        sys.exit(2)    
           
    filepath = args.filepath or ''
    filename = args.filename or ''
    subfilename=args.subfilename or ''
    
    JIRASERVICE = args.service or ''
    JIRAPROJECT = args.project or ''
    PSWD= args.password or ''
    USER= args.user or ''
    ATTACHDIR=args.attachemnts or ''
    
    # quick old-school way to check needed parameters
    if (filepath=='' or  filename=='' or JIRASERVICE=='' or  JIRAPROJECT==''  or PSWD=='' or USER=='' or subfilename=='' or ATTACHDIR=='' ):
        parser.print_help()
        sys.exit(2)
        


    Parse(filepath, filename,JIRASERVICE,JIRAPROJECT,PSWD,USER,subfilename,ATTACHDIR)


############################################################################################################################################
# Parse excel and create dictionary of
# 1) Jira main issue data
# 2) Jira subtask(s) (remark(s)) data for main issue
# 3) Info of attachment for main issue (to be added using inhouse tooling
#  
#NOTE: Uses hardcoded sheet/column value

def Parse(filepath, filename,JIRASERVICE,JIRAPROJECT,PSWD,USER,subfilename,ATTACHDIR):
    
    #false skips issue creation and other jira operations
    PROD=True #False
    ATTACHMENTS=False
    ENV="DEV" # or "PROD", sets the custom fields
    
    
    
    logging.info ("Filepath: %s     Filename:%s" %(filepath ,filename))
    files=filepath+"/"+filename
    logging.info ("Excel (main issues) file:{0}".format(files))
   
    Issues=defaultdict(dict) 

    MainSheet="general_report" 
    wb= openpyxl.load_workbook(files)
    #types=type(wb)
    #logging.debug ("Type:{0}".format(types))
    #sheets=wb.get_sheet_names()
    #logging.debug ("Sheets:{0}".format(sheets))
    CurrentSheet=wb[MainSheet] 
    #logging.debug ("CurrentSheet:{0}".format(CurrentSheet))
    #logging.debug ("First row:{0}".format(CurrentSheet['A4'].value))


    #subtasks
    logging.info ("Subtasks ---> Filepath: %s     Filename:%s" %(filepath ,subfilename))
    subfiles=filepath+"/"+subfilename
    logging.info ("Subtasks file:{0}".format(subfiles))
   
    
    SubMainSheet="general_report" 
    subwb= openpyxl.load_workbook(subfiles)
    #types=type(wb)
    #logging.debug ("Type:{0}".format(types))
    #sheets=wb.get_sheet_names()
    #logging.debug ("Sheets:{0}".format(sheets))
    SubCurrentSheet=subwb[SubMainSheet] 
    #logging.debug ("CurrentSheet:{0}".format(CurrentSheet))
    #logging.debug ("First row:{0}".format(CurrentSheet['A4'].value))



    ########################################
    #CONFIGURATIONS AND EXCEL COLUMN MAPPINGS, both main and subtask excel
    # If now subtask description, it is sma as main task one (excample C column)
    DATASTARTSROW=5 # data section starting line MAIN TASKS EXCEL
    DATASTARTSROWSUB=5 # data section starting line SUB TASKS EXCEL
    B=2 #Key (inspectionnumber NW)
    SUB_B=2 
    C=3 #SUMMARY
    SUB_C=3
    D=4 #Issue Type NW (oroginal)
    SUB_D=4
    E=5 #Issue Type
    SUB_E=5
    F=6 #Status NW (orginal status) 
    SUB_F=6
    G=7 #Status(manually mapped status to current system)
    SUB_G=7
    H=8 # Priority
    SUB_H=8 #REporter NW
    I=9 #  Responsible NW (orignal Responsible)
    SUB_I=9 # Created
    J=10 # Responsbile as a Jira user
    SUB_J=10 # Description
    K=11 #Inspection date, format: 1.11.2018  0:00:00    system number, subtasks excel
    SUB_K=11 # Ship Number   
    L=12 #ShipNumber 
    SUB_L=12 #System Number NW
    M=13 #System Number NW (original one)
    SUB_M=13 #Performer
    N=14 #System   can be not set
    SUB_N=14 #Responsible NW
    SUB_O=15 #Assignee (jira username)
    #SUB_P=16 #Assignee (jira username)
    Q=17 #Performer NW
    # P=16 #PerformerNW
    #Q=17 #Performer, subtask excel
    #R=18 #Responsible ,subtask excel
    #U=20 #Responsible Phone Number --> Not taken, field just exists in Jira
    
    SUB_R=18  #Ijnspection Data and Time
    SUB_S=19  #Department NW
    SUB_T=20 # Department
    SUB_U=21 #BlockNW
    V=22 #DepartmentNW  (original)
    SUB_V=22 #Deck NW
    W=23 #Department 
    #SUB_W=23 # Deck NW
    X=24 # Topology  --> add to description
    Y=25 # Area
    Z=26 #Surveyor
    AA=27 #DeckNW
    AB=28 #Block NW
    AC=29 #Firezone NW
    
    

  
    
    ##############################################################################################
    #Go through main excel sheet for main issue keys (and contents findings)
    # Create dictionary structure (remarks)
    # NOTE: Uses hardcoded sheet/column values
    # NOTE: As this handles first sheet, using used row/cell reading (buggy, works only for first sheet) 
    #
    i=DATASTARTSROW # brute force row indexing
    ENDROW=(CurrentSheet.max_row) # to prvent off-by-one in the end of sheet, also excel needs deleting of empty end line1
    #print "ENDROW:{0}".format(ENDROW)
    #sys.exit(1)
    for row in CurrentSheet[('B{}:B{}'.format(DATASTARTSROW,ENDROW))]:  # go trough all column B (KEY) rows
        for mycell in row:
            KEY=mycell.value
            #print "ROW:{0} Original ID:{1}".format(i,mycell.value)
            #print "KEY:{0}".format(KEY)
            Issues[KEY]={} # add to dictionary as master key (KEY)
            
            #Just hardocode operations, POC is one off (actually second time this tool variant is used......)

           
            SUMMARY=(CurrentSheet.cell(row=i, column=C).value)
            if not SUMMARY:
                SUMMARY="Summary for this task has not been defined"
            Issues[KEY]["SUMMARY"] = SUMMARY
            
            ISSUE_TYPENW=(CurrentSheet.cell(row=i, column=D).value)
            Issues[KEY]["ISSUE_TYPENW"] = ISSUE_TYPENW
            
            ISSUE_TYPE=(CurrentSheet.cell(row=i, column=E).value)
            Issues[KEY]["ISSUE_TYPE"] = ISSUE_TYPE
            
            STATUSNW=(CurrentSheet.cell(row=i, column=F).value)
            Issues[KEY]["STATUSNW"] = STATUSNW
            
          
            
            STATUS=(CurrentSheet.cell(row=i, column=G).value)
            Issues[KEY]["STATUS"] = STATUS
            
            PRIORITY=(CurrentSheet.cell(row=i, column=H).value)
            if not PRIORITY:
                SUMMARY="Major"  # force set, all should be major
            Issues[KEY]["PRIORITY"] = PRIORITY
            
            RESPONSIBLENW=(CurrentSheet.cell(row=i, column=I).value)
            Issues[KEY]["RESPONSIBLENW"] = RESPONSIBLENW
            
            
            RESPONSIBLE=(CurrentSheet.cell(row=i, column=J).value)
            Issues[KEY]["RESPONSIBLE"] = RESPONSIBLE
            
            CREATED=(CurrentSheet.cell(row=i, column=K).value) #Inspection date
            # ISO 8601 conversion to Exceli time
            time2=CREATED.strftime("%Y-%m-%dT%H:%M:%S.000-0300")  #-0300 is UTC delta to Finland, 000 just keeps Jira happy
            #print "Original date format:{0}".format(CREATED)
            #print "CREATED ISOFORMAT TIME2:{0}".format(time2)
            CREATED=time2
            INSPECTED=CREATED # just reusing value
            Issues[KEY]["INSPECTED"] = INSPECTED
            
            
            SHIPNUMBER=(CurrentSheet.cell(row=i, column=L).value)
            Issues[KEY]["SHIPNUMBER"] = SHIPNUMBER
            
            #PERFORMER=(CurrentSheet.cell(row=i, column=P).value)
            #Issues[KEY]["PERFORMER"] = PERFORMER.encode('utf-8')
            
              
            #RESPHONE=(CurrentSheet.cell(row=i, column=U).value)
            #Issues[KEY]["RESPHONE"] = RESPHONE
            
            SYSTEMNUMBERNW=(CurrentSheet.cell(row=i, column=M).value)
            Issues[KEY]["SYSTEMNUMBERNW"] = SYSTEMNUMBERNW
            
            SYSTEM=(CurrentSheet.cell(row=i, column=N).value)
            Issues[KEY]["SYSTEM"] = SYSTEM
            
            
            PERFORMERNW=(CurrentSheet.cell(row=i, column=Q).value)
            Issues[KEY]["PERFORMERNW"] = PERFORMERNW
            
            DEPARTMENTNW=(CurrentSheet.cell(row=i, column=V).value)
            Issues[KEY]["DEPARTMENTNW"] = DEPARTMENTNW
            
            DEPARTMENT=(CurrentSheet.cell(row=i, column=W).value)
            Issues[KEY]["DEPARTMENT"] = DEPARTMENT
            
                
            TOPOLOGY=(CurrentSheet.cell(row=i, column=X).value)
            if TOPOLOGY:
                DESCRIPTION="Topology ---->  "+TOPOLOGY
            Issues[KEY]["DESCRIPTION"] = DESCRIPTION    
            Issues[KEY]["TOPOLOGY"] = TOPOLOGY
            
            
            AREA=(CurrentSheet.cell(row=i, column=Y).value)
            Issues[KEY]["AREA"] = AREA
            
            SURVEYOR=(CurrentSheet.cell(row=i, column=Z).value)
            Issues[KEY]["SURVEYOR"] = SURVEYOR
            
            DECKNW=(CurrentSheet.cell(row=i, column=AA).value)
            Issues[KEY]["DECKNW"] = DECKNW
            
            BLOCKNW=(CurrentSheet.cell(row=i, column=AB).value)
            Issues[KEY]["BLOCKNW"] = BLOCKNW
            
            FIREZONENW=(CurrentSheet.cell(row=i, column=AC).value)
            Issues[KEY]["FIREZONENW"] = FIREZONENW
            
            
            
            #Create sub dictionary for possible subtasks (to be used later)
            Issues[KEY]["REMARKS"]={}
            
            #logging.debug("---------------------------------------------------")
            i=i+1
            
            # ???? key=KEY
            #HandleAttachemnts(filepath,key,ATTACHDIR)
            
    #print Issues
    #print Issues.items() 
    
    #print(json.dumps(Issues, indent=4, sort_keys=True))
    
    
    #key=18503 # check if this key exists
    #if key in Issues:
    #    print "EXISTS"
    #else:
    #    print "NOT THERE"
    #for key, value in Issues.iteritems() :
    #    print key, value


    #print "EXITNG NOW!"
    #sys.exit(5)
    

    ############################################################################################################################
    # Check any remarks (subtasks) for main issue
    # NOTE: Uses hardcoded sheet/column values
    #
    #removed currently dfue excel changes

    
    print "Checking all subtasks now"
    print "Subtasks file:{0}".format(subfilename)

    
    i=DATASTARTSROWSUB # brute force row indexing
    for row in SubCurrentSheet[('B{}:B{}'.format(DATASTARTSROWSUB,SubCurrentSheet.max_row))]:  # go trough all column B (KEY) rows
        for submycell in row:
            PARENTKEY=submycell.value
            #logging.debug("SUBROW:{0} Original PARENT ID:{1}".format(i,PARENTKEY))
            #Issues[KEY]={} # add to dictionary as master key (KEY)
            
            #Just hardocode operations, POC is one off

            if PARENTKEY in Issues:
                logging.debug( "Subtask has a known parent {0}".format(PARENTKEY))
                #REMARKKEY=SubCurrentSheet['J{0}'.format(i)].value  # column J holds Task-ID NW
                REMARKKEY=(SubCurrentSheet.cell(row=i, column=B).value) #parent key value
                REMARKKEY=str(REMARKKEY)+"_"+str(i)  # add _ROWNUBER to create really unique key 
                #print "CREATED REMARKKEY:{0}".format(REMARKKEY)
                #Issues[KEY]["REMARKS"]={}
                Issues[PARENTKEY]["REMARKS"][REMARKKEY] = {}
                
                
                # Just hardcode operattions, POC is one off
                #DECK=SubCurrentSheet['AA{0}'.format(i)].value  # column AA holds DECK
                SUBSUMMARY=(SubCurrentSheet.cell(row=i, column=SUB_C).value)
                Issues[PARENTKEY]["REMARKS"][REMARKKEY]["SUMMARY"] = SUBSUMMARY
                
                SUBISSUE_TYPENW=(SubCurrentSheet.cell(row=i, column=SUB_D).value)
                Issues[PARENTKEY]["REMARKS"][REMARKKEY]["ISSUE_TYPENW"] = SUBISSUE_TYPENW
                
                SUBISSUE_TYPE=(SubCurrentSheet.cell(row=i, column=SUB_E).value)
                Issues[PARENTKEY]["REMARKS"][REMARKKEY]["ISSUE_TYPE"] = SUBISSUE_TYPE
                
                SUBSTATUSNW=(SubCurrentSheet.cell(row=i, column=SUB_F).value)
                Issues[PARENTKEY]["REMARKS"][REMARKKEY]["STATUSNW"] = SUBSTATUSNW
                
                SUBSTATUS=(SubCurrentSheet.cell(row=i, column=SUB_G).value)
                Issues[PARENTKEY]["REMARKS"][REMARKKEY]["STATUS"] = SUBSTATUS
                
                
                SUBREPORTERNW=(SubCurrentSheet.cell(row=i, column=SUB_H).value)
                Issues[PARENTKEY]["REMARKS"][REMARKKEY]["REPORTERNW"] = SUBREPORTERNW
                
                SUBCREATED=(SubCurrentSheet.cell(row=i, column=SUB_I).value) #Inspection date
                # ISO 8601 conversion to Exceli time
                subtime2=SUBCREATED.strftime("%Y-%m-%dT%H:%M:%S.000-0300")  #-0300 is UTC delta to Finland, 000 just keeps Jira happy
                #print "CREATED SUBTASK ISOFORMAT TIME2:{0}".format(subtime2)
                SUBCREATED=subtime2
                Issues[PARENTKEY]["REMARKS"][REMARKKEY]["SUBCREATED"] = SUBCREATED
                
                SUBDESCRIPTION=(SubCurrentSheet.cell(row=i, column=SUB_J).value)
                Issues[PARENTKEY]["REMARKS"][REMARKKEY]["DESCRIPTION"] = SUBDESCRIPTION
                
                SUBSHIPNUMBER=(SubCurrentSheet.cell(row=i, column=SUB_K).value)
                Issues[PARENTKEY]["REMARKS"][REMARKKEY]["SHIPNUMBER"] = SUBSHIPNUMBER
                
                SUBSYSTEMNUMBERNW=(SubCurrentSheet.cell(row=i, column=SUB_L).value)
                Issues[PARENTKEY]["REMARKS"][REMARKKEY]["SYSTEMNUMBERNW"] = SUBSYSTEMNUMBERNW
                
                SUBPERFORMER=(SubCurrentSheet.cell(row=i, column=SUB_M).value)
                Issues[PARENTKEY]["REMARKS"][REMARKKEY]["PERFORMER"] = SUBPERFORMER
                
                SUBRESPONSIBLENW=(SubCurrentSheet.cell(row=i, column=SUB_N).value)
                Issues[PARENTKEY]["REMARKS"][REMARKKEY]["RESPONSIBLENW"] = SUBRESPONSIBLENW
                
                SUBASSIGNEE=(SubCurrentSheet.cell(row=i, column=SUB_O).value)
                Issues[PARENTKEY]["REMARKS"][REMARKKEY]["ASSIGNEE"] = SUBASSIGNEE
           
                SUBINSPECTION=(SubCurrentSheet.cell(row=i, column=SUB_R).value)
                #ISO 8601 conversion to Exceli time
                #SUBINSPECTION=SUBINSPECTION.to_datetime(SUBINSPECTION)
                subtime3=SUBINSPECTION.strftime("%Y-%m-%dT%H:%M:%S.000-0300")  #-0300 is UTC delta to Finland, 000 just keeps Jira happy
                #subtime3=SUBINSPECTION.strftime("%Y-%m-%dT%H:%M:%S.000-0300")  #-0300 is UTC delta to Finland, 000 just keeps Jira happy
                
                #print "CREATED SUBTASK ISOFORMAT TIME3:{0}".format(subtime3)
                SUBINSPECTION=subtime3
                Issues[PARENTKEY]["REMARKS"][REMARKKEY]["SUBINSPECTION"] = SUBINSPECTION
           
           
                SUBDEPARTMENTNW=(SubCurrentSheet.cell(row=i, column=SUB_S).value)
                Issues[PARENTKEY]["REMARKS"][REMARKKEY]["DEPARTMENTNW"] = SUBDEPARTMENTNW
                
                SUBDEPARTMENT=(SubCurrentSheet.cell(row=i, column=SUB_T).value)
                Issues[PARENTKEY]["REMARKS"][REMARKKEY]["DEPARTMENT"] = SUBDEPARTMENT
                
                
                SUBBLOCKNW=(SubCurrentSheet.cell(row=i, column=SUB_U).value)
                Issues[PARENTKEY]["REMARKS"][REMARKKEY]["BLOCKNW"] = SUBBLOCKNW
                
                SUBDECKNW=(SubCurrentSheet.cell(row=i, column=SUB_V).value)
                Issues[PARENTKEY]["REMARKS"][REMARKKEY]["DECKNW"] = SUBDECKNW
           
             
                #SUBTASKID=REMARKKEY
            
            else:
                    print "ERROR: Unknown parent found --> originazl key: {0}".format(PARENTKEY)
            logging.debug( "---------------------------------------------------------------------------")
            i=i+1
    
    
 
    print(json.dumps(Issues, indent=4, sort_keys=True))
    
    

    #print "EXITING NOW ALL DONE"
    #sys.exit(5)

    ##########################################################################################################################
    # Create main issues
    if (PROD==True):    
        Authenticate(JIRASERVICE,PSWD,USER)
        jira=DoJIRAStuff(USER,PSWD,JIRASERVICE)
    else:
        print "Simulated execution only"

    #create main issues
    for key, value in Issues.iteritems() :
        KEYVALUE=(key,value)
        KEY=key
        print "ORIGINAL ISSUE KEY:{0}\nVALUE:{1}".format(KEY,KEYVALUE)
        REMARKS=Issues[key]["REMARKS"]
        print "REMARKS:{0}".format(REMARKS)
        ISSUETYPE=((Issues[key]["ISSUE_TYPE"]).encode('utf-8'))
        print "JIRA ISSUE_TYPE:{0}".format(ISSUETYPE) 
        ISSUETYPENW=((Issues[key]["ISSUE_TYPENW"]).encode('utf-8')) 
        print "ORIGINAL ISSUE_TYPE:{0}".format(ISSUETYPENW)  
        STATUS=Issues[key]["STATUS"]  
        print "JIRA STATUS:{0}".format(STATUS)  
        STATUSNW=Issues[key]["STATUSNW"]
        print "ORIGINAL STATUS:{0}".format(STATUSNW)  
        PRIORITY=Issues[key]["PRIORITY"]
        print "JIRA PRIORITY:{0}".format(PRIORITY)  
        RESPONSIBLENW=str(((Issues[key]["RESPONSIBLENW"]).encode('utf8')))  
        print "ORIGINAL RESPONSIBLE:{0}".format(RESPONSIBLENW)    
        RESPONSIBLE=(Issues[key]["RESPONSIBLE"])
        print "JIRA RESPONSIBLE:{0}".format(RESPONSIBLE)    
        INSPECTEDTIME= Issues[key]["INSPECTED"]
        print "ORIGINAL CREATED TIME:{0}".format(INSPECTEDTIME)
        SHIP=Issues[key]["SHIPNUMBER"]       
        print "SHIP NUMBER:{0}".format(SHIP)  

        SYSTEM= Issues[key]["SYSTEM"]
        print "SYSTEM:{0}".format(SYSTEM) 
        SYSTEMNUMBERNW= Issues[key]["SYSTEMNUMBERNW"]
        print "SYSTEMNUMBERNW:{0}".format(SYSTEMNUMBERNW) 
        PERFORMERNW=(Issues[key]["PERFORMERNW"]).encode('utf8')
        print "ORIGINAL PERFOMER:{0}".format(PERFORMERNW)   
        DEPARTMENTNW=(Issues[key]["DEPARTMENTNW"])
        print "ORIGINAL DEPARTMENT:{0}".format(DEPARTMENTNW) 
        DEPARTMENT=(Issues[key]["DEPARTMENT"])
        print "DEPARTMENT:{0}".format(DEPARTMENT) 
        DESCRIPTION=(Issues[key]["DESCRIPTION"])
        print "DESCPTION + TOPOLOGY:{0}".format(DESCRIPTION) 

        JIRASUMMARY=(Issues[key]["SUMMARY"]).encode('utf-8')          
        JIRASUMMARY=JIRASUMMARY.replace("\n", " ") # Perl used to have chomp, this was only Python way to do this
        JIRASUMMARY=JIRASUMMARY[:254] ## summary max length is 255
        print "SUMMARY:{0}".format(JIRASUMMARY)
       
        AREA=(Issues[key]["AREA"])
        print "AREA:{0}".format(AREA) 
        
        SURVEYOR=(Issues[key]["SURVEYOR"])
        print "SURVEYOR:{0}".format(SURVEYOR) 
        
        DECKNW=(Issues[key]["DECKNW"])
        print "DECKNW:{0}".format(DECKNW) 
        
        BLOCKNW=Issues[key]["BLOCKNW"]
        print "BLOCKNW:{0}".format(BLOCKNW) 
        
        FIREZONENW=str((Issues[key]["FIREZONENW"]))  # str casting needed
        print "FIREZONENW:{0}".format(FIREZONENW) 
        
     
        
    

        
        
        #IssueID="SHIP-1826" #temp ID
        if (PROD==True):
            IssueID=CreateIssue(ENV,jira,JIRAPROJECT,JIRASUMMARY,KEY,ISSUETYPE,ISSUETYPENW,STATUS,STATUSNW,PRIORITY,RESPONSIBLENW,RESPONSIBLE,INSPECTEDTIME,SHIP,SYSTEMNUMBERNW,SYSTEM,PERFORMERNW,DEPARTMENTNW,DEPARTMENT,DESCRIPTION,AREA,SURVEYOR,DECKNW,BLOCKNW,FIREZONENW)
            print "Created issue:{0}  OK".format(IssueID)
            print "-----------------------------------------------------------"
            time.sleep(0.5)
            sys.exit(1)
            #print "IssueKey:{0}".format(IssueID.key)
        else:
           print "--> SKIPPED ISSUE CREATION" 
        
        filesx=filepath+"/*{0}*".format(key)
        print "filesx:{0}".format(filesx)
        
        
        if (ATTACHMENTS==True):
            attachments=glob.glob("{0}".format(filesx))
            if (len(attachments) > 0): # if any attachment with key embedded to name found
                print "Found attachments for key:{0}".format(IssueID)
                print "Found these:{0}".format(attachments)
                for item in attachments: # add them all
                    jira.add_attachment(issue=IssueID, attachment=attachments[0])
                    print "Attachment:{0} added".format(item)
                    time.sleep(0.5)
            else:
                print "NO attachments  found for key:{0}".format(IssueID)
        else:
            print "Skipped Attachment operations"
        
        Remarks=Issues[key]["REMARKS"] # take a copy of remarks and use it
        
        print "-------------------------------------------------------------------------"
        if (PROD==True):
            PARENT=IssueID
        #create subtask(s) under one parent
        # custom ids in comments: 1) dev 2) production
        for subkey , subvalue in Remarks.iteritems():
            
            SUBKEYVALUE=(subkey,subvalue)
            SUBKEY=subkey.encode('utf-8')
            
            ParentCheck = re.search( r"(\d*)(_)(\d*)", SUBKEY) # remove unique _ROWNUJMBER identifier
            if ParentCheck:
                CurrentGroups=ParentCheck.groups()    
                #print ("Group 1: %s" % CurrentGroups[0]) 
                #print ("Group 2: %s" % CurrentGroups[1]) 
                SUBPARENTKEY=CurrentGroups[0] #logical key (parent original key, used to tell teh parent for this subtask), dictionary key is the subkey 
            else:
                log.error("Subtask Parent parsing failure")
            print "SUBTASK PARENT'S ORIGINAL KEY:{0}\nVALUE:{1}".format(SUBPARENTKEY,SUBKEYVALUE)
            #TODO CHECK DUPLICATE REMARKS!!
            
            SUBSUMMARY=Remarks[subkey]["SUMMARY"] 
            SUBSUMMARY=SUBSUMMARY.replace("\n", "")
            SUBSUMMARY=SUBSUMMARY[:254]    ## summary max length is 255
            SUBSUMMARY=(SUBSUMMARY.encode('utf-8')) 
            print "SUBSUMMARY:{0}".format(SUBSUMMARY)
            
            SUBISSUTYPENW=Remarks[subkey]["ISSUE_TYPENW"] 
            print "SUBISSUTYPENW:{0}".format(SUBISSUTYPENW)
            SUBISSUTYPE=Remarks[subkey]["ISSUE_TYPE"] 
            print "SUBISSUTYPE:{0}".format(SUBISSUTYPE)
            
            SUBSTATUSNW=Remarks[subkey]["STATUSNW"] 
            print "SUBSTATUSNW:{0}".format(SUBSTATUSNW)
            
            SUBSTATUS=Remarks[subkey]["STATUS"] 
            print "SUBSTATUS:{0}".format(SUBSTATUS)
            
            SUBREPORTERNW=Remarks[subkey]["REPORTERNW"].encode('utf-8') 
            print "SUBREPORTERNW:{0}".format(SUBREPORTERNW)
            
            SUBCREATED=Remarks[subkey]["SUBCREATED"] 
            print "SUBCREATED:{0}".format(SUBCREATED)
            
            SUBDESCRIPTION=Remarks[subkey]["DESCRIPTION"].encode('utf-8') 
            print "SUBDESCRIPTION:{0}".format(SUBDESCRIPTION)
            
            SUBSHIPNUMBER=Remarks[subkey]["SHIPNUMBER"] 
            print "SUBSHIPNUMBER:{0}".format(SUBSHIPNUMBER)
            
            SUBSYSTEMNUMBERNW=Remarks[subkey]["SYSTEMNUMBERNW"] 
            print "SUBSYSTEMNUMBERNW:{0}".format(SUBSYSTEMNUMBERNW)
            
            SUBPERFORMER=Remarks[subkey]["PERFORMER"].encode('utf-8') 
            print "SUBPERFORMER:{0}".format(SUBPERFORMER)
            
            SUBRESPONSIBLENW=Remarks[subkey]["RESPONSIBLENW"].encode('utf-8') 
            print "SUBRESPONSIBLENW:{0}".format(SUBRESPONSIBLENW)
            
            SUBASSIGNEE=Remarks[subkey]["ASSIGNEE"] 
            print "SUBASSIGNEE:{0}".format(SUBASSIGNEE)
            
            SUBINSPECTION=Remarks[subkey]["SUBINSPECTION"] 
            print "SUBINSPECTION:{0}".format(SUBINSPECTION)
            
            SUBDEPARTMENTNW=Remarks[subkey]["DEPARTMENTNW"] 
            print "SUBDEPARTMENTNW:{0}".format(SUBDEPARTMENTNW)
            
            SUBDEPARTMENT=Remarks[subkey]["DEPARTMENT"] 
            print "SUBDEPARTMENT:{0}".format(SUBDEPARTMENT)
            
            SUBBLOCKNW=Remarks[subkey]["BLOCKNW"] 
            print "SUBBLOCKNW:{0}".format(SUBBLOCKNW)
            
            SUBDECKNW=Remarks[subkey]["DECKNW"] 
            print "SUBDECKNW:{0}".format(SUBDECKNW)
            
            print ".................................."
            if (PROD==True):
                #SubIssueID=CreateSubTask(jira,JIRAPROJECT,SUBSUMMARY,SUBISSUTYPENW,SUBISSUTYPE,SUBSTATUSNW,SUBSTATUS,SUBREPORTERNW,SUBCREATED,SUBDESCRIPTION,SUBSHIPNUMBER,SUBSYSTEMNUMBERNW,SUBPERFORMER,SUBRESPONSIBLENW,SUBASSIGNEE,SUBINSPECTION,SUBDEPARTMENTNW,SUBDEPARTMENT,SUBBLOCKNW,SUBDECKNW)
                print "Created subtask:{0}".format(SubIssueID)
                time.sleep(0.5)
            else:
                print "Skipped subtask creation"
            
        print "*************************************************************************"
        
#############################################################################

def HandleAttachemnts(filepath,key,ATTACHDIR):
        filesx=ATTACHDIR+"/*{0}*".format(key)
        print "filesx:{0}".format(filesx)
        
        
        attachments=glob.glob("{0}".format(filesx))
        if (len(attachments) > 0): # if any attachment with key embedded to name found
            print "===> Found attachments for key:{0}".format(key)
            print "Found these:{0}".format(attachments)
            #for item in attachments: # add them all
            #    jira.add_attachment(issue=IssueID, attachment=attachments[0])
            #    print "Attachment:{0} added".format(item)
        else:
            print "NO attachments  found for original key:{0}".format(key)
        print "******************************************************************************************************************"

    
logging.debug ("--Python exiting--")
if __name__ == "__main__":
    main(sys.argv[1:]) 